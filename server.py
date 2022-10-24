from datetime import datetime
from unidecode import unidecode
import json
import openpyxl
from fastapi import FastAPI, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from typing import List
import databases
import pandas as pd
import sqlalchemy
from sqlalchemy_utils import URLType
from pydantic import BaseModel
import os
import boto3
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
from itertools import islice

load_dotenv()
DATABASE_URL = (os.environ["DATABASE"])
bucket_name = (os.environ["AWS_BUCKET_NAME"])

s3 = boto3.client(
    service_name="s3",
    aws_access_key_id=(os.environ["AWS_ACCESS_KEY_ID"]),
    aws_secret_access_key=(os.environ["AWS_SECRET_ACCESS_KEY"]),
    region_name='sa-east-1'
)

database = databases.Database(DATABASE_URL)

metadata = sqlalchemy.MetaData()

files = sqlalchemy.Table(
    "files",
    metadata,
    sqlalchemy.Column("id", sqlalchemy.Integer, primary_key=True),
    sqlalchemy.Column("name", sqlalchemy.String, nullable=False,),
    sqlalchemy.Column("file_url", URLType),
    sqlalchemy.Column("upload_at", sqlalchemy.String,
                      nullable=False, default=datetime.now())
)
engine = sqlalchemy.create_engine(
    DATABASE_URL
)
metadata.create_all(engine)


class Files(BaseModel):
    id: int
    name: str
    file_url: str
    upload_at: str


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await database.connect()


@app.on_event("shutdown")
async def shutdown():
    await database.disconnect()


@app.get("/files", response_model=List)
async def read_files():
    query = files.select().order_by("id")
    return await database.fetch_all(query)


@app.post("/file/upload")
async def upload_file(arquivo: UploadFile):
    upload_at = datetime.now()

    s3.upload_fileobj(arquivo.file, bucket_name, arquivo.filename)

    uploaded_file_url = f"https://{bucket_name}.s3.amazonaws.com/{unidecode(str(arquivo.filename))}"

    query = files.insert().values(name=str(arquivo.filename),
                                  file_url=uploaded_file_url, upload_at=str(upload_at))

    last_record_id = await database.execute(query)

    return {"name": arquivo.filename, "url_file": uploaded_file_url, "id": last_record_id, "upload_at": upload_at}


@app.put("/file/{file_id}")
async def update_file(file_id: int, arquivo: UploadFile):
    upload_at = datetime.now()

    s3.upload_fileobj(arquivo.file, bucket_name, arquivo.filename)

    uploaded_file_url = f"https://{bucket_name}.s3.amazonaws.com/{unidecode(str(arquivo.filename))}"
    query = files.update().where(files.columns.id == file_id).values(
        name=arquivo.filename, file_url=uploaded_file_url, upload_at=str(upload_at))
    await database.execute(query)
    return {"id": file_id, "name": arquivo.filename, "file_url": uploaded_file_url, "upload_at": str(upload_at)}


@app.delete("/file/delete/{file_id}")
async def update_file(file_id: int):
    query = files.delete().where(files.columns.id == file_id)
    await database.execute(query)
    return {"message": "the file with id={} deleted successfully".format(file_id)}


@app.get("/file/conciliado")
async def conciliado():
    query = "SELECT file_url FROM files ORDER BY upload_at DESC LIMIT 5"
    await database.execute(query)
    var = await database.fetch_all(query)
    vendas_cielo_url = str(tuple(var[4].values())).replace("('", "")
    vendas_cielo_url_replace = vendas_cielo_url.replace("',)", "")
    vendas_sig_url = str(tuple(var[3].values())).replace("('", "")
    vendas_sig_url_replace = vendas_sig_url.replace("',)", "")
    recebimentos_cielo_url = str(tuple(var[2].values())).replace("('", "")
    recebimentos_cielo_url_replace = recebimentos_cielo_url.replace("',)", "")
    recebimentos_sig_url = str(tuple(var[1].values())).replace("('", "")
    recebimentos_sig_url_replace = recebimentos_sig_url.replace("',)", "")
    mxm_url = str(tuple(var[0].values())).replace("('", "")
    mxm_url_replace = mxm_url.replace("',)", "")

    df_vendas_cielo = pd.read_excel(vendas_cielo_url_replace, usecols=[
        "Código de autorização", "Valor da venda"])
    df_vendas_sig = pd.read_excel(vendas_sig_url_replace, usecols=[
        "Data da Venda", "Aut. de Venda", "Valor Proporcional"])
    df_recebimentos_cielo = pd.read_excel(recebimentos_cielo_url_replace, usecols=[
        "Data de pagamento", "Código de autorização", "Valor bruto"])
    df_recebimentos_sig = pd.read_excel(recebimentos_sig_url_replace, usecols=[
        "Data do Recebimento", "Aut. de Venda", "Valor Proporcional"])
    df_mxm = pd.read_excel(mxm_url_replace, usecols=[
                           'Data', 'Histórico', 'Débito', 'Crédito'], skipfooter=1)

    grup_vendas_cielo = df_vendas_cielo.groupby(
        pd.Grouper(key='Código de autorização')).sum()
    grup_vendas_sig = df_vendas_sig.groupby(
        pd.Grouper(key='Aut. de Venda')).sum()
    grup_recebimentos_cielo = df_recebimentos_cielo.groupby(
        pd.Grouper(key='Código de autorização')).sum()
    grup_recebimentos_sig = df_recebimentos_sig.groupby(
        pd.Grouper(key='Aut. de Venda')).sum()
    grup_recebimentos_sig2 = df_recebimentos_sig.groupby(
        pd.Grouper(key='Data do Recebimento')).sum()
    grup_vendas_sig2 = df_vendas_sig.groupby(
        pd.Grouper(key='Data da Venda')).sum()
    grup_mxm = df_mxm.groupby(pd.Grouper(key="Data")).sum()

    vendas_cieloXsig = pd.merge(pd.DataFrame(
        grup_vendas_cielo), pd.DataFrame(grup_vendas_sig), left_on="Código de autorização", right_on="Aut. de Venda", right_index=True)

    recebimentos_cieloXsig = pd.merge(pd.DataFrame(grup_recebimentos_cielo), pd.DataFrame(
        grup_recebimentos_sig), left_on="Código de autorização", right_on="Aut. de Venda", right_index=True)

    razao_contabilXrecebimentos_sig = pd.merge(pd.DataFrame(grup_recebimentos_sig2), pd.DataFrame(
        grup_mxm), left_on="Data do Recebimento", right_on="Data", right_index=True)

    razao_contabilXvendas_sig = pd.merge(pd.DataFrame(grup_vendas_sig2), pd.DataFrame(
        grup_mxm), left_on="Data da Venda", right_on="Data", right_index=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df_vendas_cielo, index=True, header=True):
        ws.append(r)

    wb.create_sheet('diferencas_vendas_cieloxsig')
    diferencas_vendas_cieloxsig = wb['diferencas_vendas_cieloxsig']
    diferencas_vendas_cieloxsig.append(
        ['id', 'aut_pagamento', 'valor_cielo', 'valor_sig', 'diferenca'])

    for index, row in vendas_cieloXsig.iterrows():
        v_cielo = row["Valor da venda"]
        v_sig = row["Valor Proporcional"]
        if round(v_cielo) != round(v_sig):
            diferenca = round(v_cielo, 2) - round(v_sig, 2)
            diferencas_vendas_cieloxsig.append([str(
                index), str(
                index), v_cielo, v_sig, str(round(diferenca, 2)).replace("-", "")])

    wb.create_sheet('diferencas_recebimentos_cieloxsig')
    diferencas_recebimentos_cieloxsig = wb['diferencas_recebimentos_cieloxsig']
    diferencas_recebimentos_cieloxsig.append(
        ['id', 'aut_pagamento', 'valor_cielo', 'valor_sig', 'diferenca'])

    for index, row in recebimentos_cieloXsig.iterrows():
        r_cielo = row["Valor bruto"]
        r_sig = row["Valor Proporcional"]
        if round(r_cielo) != round(r_sig):
            diferenca = round(r_cielo, 2) - round(r_sig, 2)
            diferencas_recebimentos_cieloxsig.append([str(
                index), str(
                index), r_cielo, r_sig, str(round(diferenca, 2)).replace("-", "")])

    wb.create_sheet('diferencas_recebimentos_mxm_sig')
    diferenca_recebimentos_sigxmxm = wb['diferencas_recebimentos_mxm_sig']
    diferenca_recebimentos_sigxmxm.append(
        ['id', 'data_recebimento', 'valor_sig', 'valor_mxm', 'diferenca'])

    for index, row in razao_contabilXrecebimentos_sig.iterrows():
        v_mxm_credito = row["Crédito"]
        v_sig_recebimento = row["Valor Proporcional"]
        if round(v_sig_recebimento) != round(v_mxm_credito):
            diferenca_recebimentos = round(
                v_sig_recebimento, 2) - round(v_mxm_credito, 2)
            diferenca_recebimentos_sigxmxm.append([str(index), str(
                index), v_sig_recebimento, v_mxm_credito, str(round(diferenca_recebimentos, 2)).replace("-", "")])

    wb.create_sheet('diferencas_vendas_mxm_sig')
    diferenca_vendas_sigxmxm = wb['diferencas_vendas_mxm_sig']
    diferenca_vendas_sigxmxm.append(
        ['id', 'data_venda', 'valor_sig', 'valor_mxm', 'diferenca'])

    for index, row in razao_contabilXvendas_sig.iterrows():
        v_mxm_debito = row["Débito"]
        v_sig_venda = row["Valor Proporcional"]
        if round(v_sig_venda) != round(v_mxm_debito):
            diferenca_vendas = round(v_sig_venda, 2) - round(v_mxm_debito, 2)
            diferenca_vendas_sigxmxm.append([str(index), str(
                index), v_sig_venda, v_mxm_debito, str(round(diferenca_vendas, 2)).replace("-", "")])

    dados_vendas = wb["diferencas_vendas_cieloxsig"].values
    cols_vendas = next(dados_vendas)[1:]
    dados_vendas = list(dados_vendas)
    idx_vendas = [r[0] for r in dados_vendas]
    dados_vendas = (islice(r, 1, None) for r in dados_vendas)
    dif_vendas = pd.DataFrame(
        dados_vendas, index=idx_vendas, columns=cols_vendas).to_json(orient="records")

    dados_recebimentos = wb["diferencas_recebimentos_cieloxsig"].values
    cols_recebimentos = next(dados_recebimentos)[1:]
    dados_recebimentos = list(dados_recebimentos)
    idx_recebimentos = [r[0] for r in dados_recebimentos]
    dados_recebimentos = (islice(r, 1, None) for r in dados_recebimentos)
    dif_recebimentos = pd.DataFrame(
        dados_recebimentos, index=idx_recebimentos, columns=cols_recebimentos).to_json(orient="records")

    dados_recebimentos_sig_mxm = wb["diferencas_recebimentos_mxm_sig"].values
    cols_recebimentos_mxm_sig = next(dados_recebimentos_sig_mxm)[1:]
    dados_recebimentos_sig_mxm = list(dados_recebimentos_sig_mxm)
    idx_recebimentos_mxm_sig = [r[0] for r in dados_recebimentos_sig_mxm]
    dados_recebimentos_sig_mxm = (islice(r, 1, None)
                                  for r in dados_recebimentos_sig_mxm)
    dif_recebimentos_sig_razao_contabil = pd.DataFrame(
        dados_recebimentos_sig_mxm, index=idx_recebimentos_mxm_sig, columns=cols_recebimentos_mxm_sig).to_json(orient="records")

    dados_vendas_sig_mxm = wb["diferencas_vendas_mxm_sig"].values
    cols_vendas_mxm_sig = next(dados_vendas_sig_mxm)[1:]
    dados_vendas_sig_mxm = list(dados_vendas_sig_mxm)
    idx_vendas_sig_mxm = [r[0] for r in dados_vendas_sig_mxm]
    dados_vendas_sig_mxm = (islice(r, 1, None) for r in dados_vendas_sig_mxm)
    dif_vendas_sig_razao_contabil = pd.DataFrame(
        dados_vendas_sig_mxm, index=idx_vendas_sig_mxm, columns=cols_vendas_mxm_sig).to_json(orient="records")

    json_vendas = json.loads(str(dif_vendas).replace("\\", ""))
    json_recebimentos = json.loads(str(dif_recebimentos).replace("\\", ""))
    json_recebimentos_razao_sig = json.loads(
        str(dif_recebimentos_sig_razao_contabil).replace("\\", ""))
    json_vendas_razao_sig = json.loads(
        str(dif_vendas_sig_razao_contabil).replace("\\", ""))

    json_ob = [{"dif_vendas_cielo_sig": json_vendas, "dif_recebimentos_cielo_sig":
               json_recebimentos, "dif_recebimentos_sig_mxm": json_recebimentos_razao_sig, "dif_vendas_sig_mxm": json_vendas_razao_sig}]

    # response_file =

    # dif_vendas_df = pd.DataFrame(
    #     dados_vendas, index=idx_vendas, columns=cols_vendas)
    # dif_recebimentos_df = pd.DataFrame(
    #     dados_recebimentos, index=idx_recebimentos, columns=cols_recebimentos)
    # dif_razao_contabil_df = pd.DataFrame(
    #     dados_recebimentos_sig_mxm, index=idx_mxm, columns=cols_mxm)

    # df_concat = pd.concat(
    #     [dif_vendas_df, dif_recebimentos_df, dif_razao_contabil_df])
    # file_name = "diferencas.csv"
    # csv_buffer = StringIO()
    # s3_resource = boto3.resource('s3')

    # pd.DataFrame(df_concat).to_csv(csv_buffer)

    # s3_resource.Object(bucket_name, file_name).put(Body=csv_buffer.getvalue())
    # s3.put_object(
    #     ACL="public",
    #     Body=excel_buffer.getvalue(),
    #     Bucket=bucket_name,
    #     Key=file_name
    # )
    # s3.upload_fileobj(hashlib.sha256(excel_buffer).hexdigest(), bucket_name, file_name)
    # filename, storage_options = s3.upload_fileobj(json_ob, bucket_name, filename))

    # uploaded_file_url = f"https://{bucket_name}.s3.amazonaws.com/{str(arquivo.filename)}"
    return json_ob
